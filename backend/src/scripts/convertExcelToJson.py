#!/usr/bin/env python3
"""Convert the dosage tables Excel workbook to JSON for Node.js consumption.

Bypasses the ExcelJS 31-character sheet name limitation.

Usage: python3 src/scripts/convertExcelToJson.py
"""
import openpyxl
import json
import os

DATA_DIR = os.path.join(os.path.dirname(__file__), '..', '..', 'data')
XLSX_PATH = os.path.join(DATA_DIR, 'Peptide_dosaing_tables_replaced_completed (1).xlsx')
JSON_PATH = os.path.join(DATA_DIR, 'dosage_tables.json')

wb = openpyxl.load_workbook(XLSX_PATH)
result = {"index": [], "sheets": {}}

# Parse Index sheet
ws_index = wb["Index"]
headers = None
for i, row in enumerate(ws_index.iter_rows(values_only=True)):
    if i == 0:
        headers = [str(v or "").strip() for v in row]
        continue
    obj = {}
    for j, h in enumerate(headers):
        if h and j < len(row):
            val = row[j]
            obj[h] = str(val).strip() if val is not None else ""
    title = obj.get("Protocol Title (Column A)", "")
    if title:
        result["index"].append(obj)

# Parse individual dosage sheets
for sheet_name in wb.sheetnames:
    if sheet_name == "Index":
        continue
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        cells = []
        for v in row:
            cells.append(str(v).strip() if v is not None else None)
        rows.append(cells)
    result["sheets"][sheet_name] = rows

with open(JSON_PATH, "w", encoding="utf-8") as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

n_index = len(result["index"])
n_sheets = len(result["sheets"])
print(f"Converted {n_index} index entries and {n_sheets} sheets to JSON")
print(f"Output: {JSON_PATH}")
